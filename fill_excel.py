from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime
import os
import re


def fill_excel_multi(all_data):

    os.makedirs("outputs", exist_ok=True)

    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    for i, data in enumerate(all_data):
        # Create a new sheet for each bill
        sheet_name = f"Bill_{i+1}"
        if data.get("consumer_name"):
            # Clean sheet name (max 31 chars, no special chars)
            sheet_name = re.sub(r'[\\*?:/\[\]]', '', data.get("consumer_name"))[:30]
        
        ws = wb.create_sheet(title=sheet_name)

        # -------------------------------------------------
        # STYLES
        # -------------------------------------------------

        bold = Font(bold=True)

        header_fill = PatternFill(
            start_color="FFA500",
            end_color="FFA500",
            fill_type="solid"
        )

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # -------------------------------------------------
        # CUSTOMER DETAILS (Columns B and D)
        # -------------------------------------------------

        # Row 1: Consumer Name
        ws["B1"] = "Consumer Name"
        ws["D1"] = data.get("consumer_name", "")
        ws["B1"].font = bold
        ws["B1"].fill = header_fill
        ws["B1"].border = border
        ws["D1"].border = border

        # Row 2: Consumer No
        ws["B2"] = "Consumer No"
        ws["D2"] = str(data.get("consumer_number", ""))
        ws["B2"].font = bold
        ws["B2"].fill = header_fill
        ws["B2"].border = border
        ws["D2"].border = border
        ws["D2"].number_format = '@'


        # Row 3: Fixed Charges
        ws["B3"] = "Fixed Charges"
        ws["D3"] = data.get("fixed_charges", "130")
        ws["B3"].font = bold
        ws["B3"].fill = header_fill
        ws["B3"].border = border
        ws["D3"].border = border

        # Row 4: Sanct. Load (kW)
        # Row 4: Sanct. Load (kW)

        load_value = data.get("load_kw", "")

        if not load_value:
            load_value = "3"

        ws["B4"] = "Sanct. Load (kW)"
        ws["D4"] = f"{load_value} KW"
        ws["B4"].font = bold
        ws["B4"].fill = header_fill
        ws["B4"].border = border
        ws["D4"].border = border

        # Row 5: Connection Type
        # Row 5: Connection Type

        tariff_value = data.get("tariff", "")

        if tariff_value == "A50" or not tariff_value:
            tariff_value = "90/LT I Res 1-Phase"

        ws["B5"] = "Connection Type"
        ws["D5"] = tariff_value
        ws["B5"].font = bold
        ws["B5"].fill = header_fill
        ws["B5"].border = border
        ws["D5"].border = border

        # Row 6: Contract Demand
        ws["B6"] = "Contract Demand (KVA) :"
        ws["B6"].font = bold

        # Row 7: Solar Panel Used
        ws["B7"] = "Solar Pannel used"
        ws["C7"] = 600
        ws["B7"].font = bold
        ws["B7"].fill = header_fill
        ws["B7"].border = border
        ws["C7"].fill = yellow_fill
        ws["C7"].border = border

        # -------------------------------------------------
        # TABLE HEADER
        # -------------------------------------------------

        start_row = 9
        headers = ["Sr.No", "Month", "Units", "Bill Amount", "Unit Cost"]
        cols = [2, 3, 4, 5, 6] # B, C, D, E, F

        for col, header in zip(cols, headers):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = bold
            cell.fill = header_fill
            cell.border = border

        # -------------------------------------------------
        # MONTHLY HISTORY DATA
        # -------------------------------------------------

        monthly_history = data.get("monthly_history", [])
        total_units = 0

        # Mappings for full month names
        month_map = {

            "JAN 2024": "January 2024",
            "FEB 2024": "February 2024",
            "MAR 2024": "March 2024",
            "APR 2024": "April 2024",
            "MAY 2024": "May 2024",
            "JUN 2024": "June 2024",
            "JUL 2024": "July 2024",
            "AUG 2024": "August 2024",
            "SEP 2024": "September 2024",
            "OCT 2024": "October 2024",
            "NOV 2024": "November 2024",
            "DEC 2024": "December 2024",

            "Mar-2024": "March 2024",
            "Feb-2024": "February 2024",
            "Jan-2024": "January 2024",
            "Dec-2023": "December 2023",
            "Nov-2023": "November 2023",
            "Oct-2023": "October 2023",

            "JAN 2025": "January 2025",
            "FEB 2025": "February 2025",
            "MAR 2025": "March 2025",
            "APR 2025": "April 2025",
            "MAY 2025": "May 2025",
            "JUN 2025": "June 2025",
            "JUL 2025": "July 2025",
            "AUG 2025": "August 2025",
            "SEP 2025": "September 2025",
            "OCT 2025": "October 2025",
            "NOV 2025": "November 2025",
            "DEC 2025": "December 2025",

            "JAN 2026": "January 2026",
            "FEB 2026": "February 2026",
            "MAR 2026": "March 2026"
        }
        for index in range(13): # Show 13 rows like in original
            row_idx = start_row + index + 1
            ws.cell(row=row_idx, column=2).value = index + 2
                    
            if index < len(monthly_history):

                item = monthly_history[index]

                # Get month safely
                m_short = str(
                    item.get("month", "")
                ).strip()

                # Convert month format
                month_value = month_map.get(
                    m_short,
                    m_short
                )

                # Write month to Excel
                ws.cell(
                    row=row_idx,
                    column=3
                ).value = month_value

                # Units
                units_value = item.get(
                    "units",
                    0
                )

                try:
                    units_value = int(units_value)

                except:
                    units_value = 0

                # Write units
                ws.cell(
                    row=row_idx,
                    column=4
                ).value = units_value

                # Total calculation
                total_units += units_value

                # DEBUG
                print(
                    f"Excel Row => {month_value} : {units_value}"
                )
            
            for col in range(2, 7):
                ws.cell(row=row_idx, column=col).border = border

        # -------------------------------------------------
        # CALCULATIONS
        # -------------------------------------------------

        calc_row = start_row + 14 # Fixed position for calculations

        avg_units = total_units / len(monthly_history) if len(monthly_history) > 0 else 0
        kw = avg_units / 106
        solar_panels = kw / 0.6
        solar_capacity = round(solar_panels * 0.7, 1)
        num_panels = round(solar_capacity / 0.6)

        calculations = [
            ("Average", round(avg_units, 2)),
            ("kW", round(kw, 2)),
            ("Solar Panels", round(solar_panels, 2)),
            ("Solar capacity", solar_capacity),
            ("Number of Panels", num_panels),
        ]

        # Formatting colors for bottom rows
        solar_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid") # Orange-ish

        for label, value in calculations:
            ws.cell(row=calc_row, column=3).value = label
            ws.cell(row=calc_row, column=4).value = value
            ws.cell(row=calc_row, column=3).font = bold
            ws.cell(row=calc_row, column=3).border = border
            ws.cell(row=calc_row, column=4).border = border
            
            if label == "Solar capacity":
                ws.cell(row=calc_row, column=3).fill = solar_fill
                ws.cell(row=calc_row, column=4).fill = yellow_fill
            if label == "Number of Panels":
                ws.cell(row=calc_row, column=3).fill = green_fill
                ws.cell(row=calc_row, column=4).fill = green_fill

            calc_row += 1

        # Bottom summary
        ws.cell(row=calc_row+2, column=3).value = "Total solar capacity"
        ws.cell(row=calc_row+2, column=4).value = solar_capacity * 2
        ws.cell(row=calc_row+3, column=3).value = "Number of solar panels"
        ws.cell(row=calc_row+3, column=4).value = num_panels * 2

        # Column widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15


    # -------------------------------------------------
    # SAVE
    # -------------------------------------------------

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"outputs/all_bills_{timestamp}.xlsx"

    wb.save(output_file)

    print(f"\nExcel saved: {output_file}")

    return output_file